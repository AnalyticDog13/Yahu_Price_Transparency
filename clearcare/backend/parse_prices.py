#!/usr/bin/env python3
"""
ClearCare data pipeline.
Reads all registered hospital CSVs, extracts imaging/diagnostic procedures
by CPT code, and writes prices.csv + prices.db from scratch.

Usage:
    python3 parse_prices.py          # rebuild everything from scratch

To add a single new hospital without rebuilding from scratch, use:
    python3 add_hospital.py --help
"""

import csv
import os
import re
import sqlite3
import time

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "hospital-price-data")
OUT_CSV  = os.path.join(os.path.dirname(__file__), "data", "prices.csv")
OUT_DB   = os.path.join(os.path.dirname(__file__), "data", "prices.db")

# CPT code → specific procedure name
PROCEDURE_CODES = {
    "70551": "MRI Brain (w/o contrast)",
    "70552": "MRI Brain (w/ contrast)",
    "70553": "MRI Brain (w/o & w/ contrast)",
    "72141": "MRI Cervical Spine (w/o contrast)",
    "72142": "MRI Cervical Spine (w/ contrast)",
    "72146": "MRI Thoracic Spine (w/o contrast)",
    "72147": "MRI Thoracic Spine (w/ contrast)",
    "72148": "MRI Lumbar Spine (w/o contrast)",
    "72149": "MRI Lumbar Spine (w/ contrast)",
    "72156": "MRI Cervical Spine (w/o & w/ contrast)",
    "72157": "MRI Thoracic Spine (w/o & w/ contrast)",
    "72158": "MRI Lumbar Spine (w/o & w/ contrast)",
    "73721": "MRI Knee/Joint (w/o contrast)",
    "73722": "MRI Knee/Joint (w/ contrast)",
    "73723": "MRI Knee/Joint (w/o & w/ contrast)",
    "74181": "MRI Abdomen (w/o contrast)",
    "74182": "MRI Abdomen (w/ contrast)",
    "74183": "MRI Abdomen (w/o & w/ contrast)",
    "70450": "CT Head/Brain (w/o contrast)",
    "70460": "CT Head/Brain (w/ contrast)",
    "70470": "CT Head/Brain (w/o & w/ contrast)",
    "71250": "CT Chest (w/o contrast)",
    "71260": "CT Chest (w/ contrast)",
    "71270": "CT Chest (w/o & w/ contrast)",
    "74176": "CT Abdomen & Pelvis (w/o contrast)",
    "74177": "CT Abdomen & Pelvis (w/ contrast)",
    "74178": "CT Abdomen & Pelvis (w/o & w/ contrast)",
    "76700": "Ultrasound Abdomen (complete)",
    "76705": "Ultrasound Abdomen (limited)",
    "76856": "Ultrasound Pelvis (complete)",
    "76857": "Ultrasound Pelvis (limited)",
    "71045": "Chest X-Ray (1 view)",
    "71046": "Chest X-Ray (2 views)",
    "85025": "CBC Blood Panel (automated w/ diff)",
    "85027": "CBC Blood Panel (automated)",
}

# CPT code → broad UI category
PROCEDURE_CATEGORIES = {
    "70551": "MRI Brain",            "70552": "MRI Brain",            "70553": "MRI Brain",
    "72141": "MRI Spine (Cervical)", "72142": "MRI Spine (Cervical)", "72156": "MRI Spine (Cervical)",
    "72146": "MRI Spine (Thoracic)", "72147": "MRI Spine (Thoracic)", "72157": "MRI Spine (Thoracic)",
    "72148": "MRI Spine (Lumbar)",   "72149": "MRI Spine (Lumbar)",   "72158": "MRI Spine (Lumbar)",
    "73721": "MRI Knee / Joint",     "73722": "MRI Knee / Joint",     "73723": "MRI Knee / Joint",
    "74181": "MRI Abdomen",          "74182": "MRI Abdomen",          "74183": "MRI Abdomen",
    "70450": "CT Head / Brain",      "70460": "CT Head / Brain",      "70470": "CT Head / Brain",
    "71250": "CT Chest",             "71260": "CT Chest",             "71270": "CT Chest",
    "74176": "CT Abdomen & Pelvis",  "74177": "CT Abdomen & Pelvis",  "74178": "CT Abdomen & Pelvis",
    "76700": "Ultrasound Abdomen",   "76705": "Ultrasound Abdomen",
    "76856": "Ultrasound Pelvis",    "76857": "Ultrasound Pelvis",
    "71045": "Chest X-Ray",          "71046": "Chest X-Ray",
    "85025": "CBC Blood Panel",      "85027": "CBC Blood Panel",
}

# Hospital registry — add new hospitals here, then run parse_prices.py
# (or use add_hospital.py to append a single hospital without a full rebuild)
HOSPITAL_FILES = [
    {
        "filename":         "510064326_St.-Francis-Hospital-Inc_standardcharges.csv",
        "hospital_name":    "St. Francis Hospital",
        "hospital_city":    "Wilmington",
        "hospital_state":   "DE",
        "hospital_address": "701 N. Clayton St",
    },
    {
        "filename":         "23-1529076_riddle-memorial-hospital_standardcharges.csv",
        "hospital_name":    "Riddle Memorial Hospital",
        "hospital_city":    "Media",
        "hospital_state":   "PA",
        "hospital_address": "1068 West Baltimore Pike",
    },
    {
        "filename":         "232829095_jefferson-methodist-hospital_standardcharges.csv",
        "hospital_name":    "Jefferson Methodist Hospital",
        "hospital_city":    "Philadelphia",
        "hospital_state":   "PA",
        "hospital_address": "2301 S Broad Street",
    },
    {
        "filename":         "231352160_the-bryn-mawr-hospital_standardcharges.xlsx",
        "hospital_name":    "Bryn Mawr Hospital",
        "hospital_city":    "Bryn Mawr",
        "hospital_state":   "PA",
        "hospital_address": "130 South Bryn Mawr Ave",
        "xlsx_sheet":       "in",
    },
]

CSV_COLUMNS = [
    "hospital_name", "hospital_city", "hospital_state", "hospital_address",
    "procedure_category", "procedure_name", "cpt_code",
    "payer", "plan",
    "negotiated_dollar", "discounted_cash", "gross_charge",
    "setting", "billing_class",
]


def safe_float(val):
    try:
        f = float(val)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _process_row(hospital: dict, row: dict) -> dict | None:
    """
    Shared row-processing logic for both CSV and XLSX parsers.
    Returns a normalized dict or None if the row should be skipped.
    """
    cpt_code = None
    for slot in range(1, 6):
        code  = str(row.get(f"code|{slot}") or "").strip()
        ctype = str(row.get(f"code|{slot}|type") or "").strip().upper()
        if ctype in ("CPT", "HCPCS") and code in PROCEDURE_CODES:
            cpt_code = code
            break

    if not cpt_code:
        return None

    negotiated = safe_float(row.get("standard_charge|negotiated_dollar"))
    cash       = safe_float(row.get("standard_charge|discounted_cash"))
    gross      = safe_float(row.get("standard_charge|gross"))

    if negotiated is None and cash is None:
        return None

    payer = re.sub(r"\s+", " ", str(row.get("payer_name") or "").strip().upper())

    return {
        "hospital_name":      hospital["hospital_name"],
        "hospital_city":      hospital["hospital_city"],
        "hospital_state":     hospital["hospital_state"],
        "hospital_address":   hospital["hospital_address"],
        "procedure_category": PROCEDURE_CATEGORIES[cpt_code],
        "procedure_name":     PROCEDURE_CODES[cpt_code],
        "cpt_code":           cpt_code,
        "payer":              payer,
        "plan":               str(row.get("plan_name") or "").strip(),
        "negotiated_dollar":  negotiated,
        "discounted_cash":    cash,
        "gross_charge":       gross,
        "setting":            str(row.get("setting") or "").strip().lower(),
        "billing_class":      str(row.get("billing_class") or "").strip().lower(),
    }


def parse_hospital_csv(hospital: dict) -> list[dict]:
    """
    Parse one hospital file (CSV or XLSX) into normalized rows.

    CSV: handles both lowercase (St. Francis, Riddle) and Title/Pascal case
         (Jefferson Methodist) column names by lowercasing headers at read time.
    XLSX: reads via openpyxl; same 2-row metadata header structure as CSV.
          Requires 'xlsx_sheet' key in hospital dict (defaults to first sheet).
    """
    path = os.path.join(DATA_DIR, hospital["filename"])
    if not os.path.exists(path):
        print(f"  SKIPPED (file not found): {hospital['filename']}")
        return []

    ext = os.path.splitext(hospital["filename"])[1].lower()
    rows = []
    t0 = time.time()
    print(f"  Parsing {hospital['hospital_name']} ({ext}) …")

    if ext == ".xlsx":
        if not XLSX_AVAILABLE:
            print("  SKIPPED — openpyxl not installed. Run: pip install openpyxl")
            return []

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_name = hospital.get("xlsx_sheet", wb.sheetnames[0])
        ws = wb[sheet_name]
        iter_rows = ws.iter_rows(values_only=True)

        next(iter_rows)  # metadata header row
        next(iter_rows)  # metadata values row
        headers = [str(h).lower() if h is not None else "" for h in next(iter_rows)]

        for i, raw_row in enumerate(iter_rows):
            row = dict(zip(headers, raw_row))
            result = _process_row(hospital, row)
            if result:
                rows.append(result)
            if i % 100_000 == 0 and i > 0:
                print(f"    … {i:,} rows scanned ({len(rows):,} matches)")

        wb.close()

    else:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            next(f)  # metadata header row
            next(f)  # metadata values row
            reader = csv.DictReader(f)
            reader.fieldnames = [h.lower() for h in reader.fieldnames]

            for i, row in enumerate(reader):
                result = _process_row(hospital, row)
                if result:
                    rows.append(result)
                if i % 500_000 == 0 and i > 0:
                    print(f"    … {i:,} rows scanned ({len(rows):,} matches)")

    print(f"  Done: {len(rows):,} rows in {time.time() - t0:.1f}s")
    return rows


def write_csv(rows: list[dict]):
    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    size_kb = os.path.getsize(OUT_CSV) / 1024
    print(f"CSV written  → {OUT_CSV}  ({size_kb:.0f} KB, {len(rows):,} rows)")


def write_db(rows: list[dict]):
    os.makedirs(os.path.dirname(OUT_DB), exist_ok=True)
    if os.path.exists(OUT_DB):
        os.remove(OUT_DB)
    con = sqlite3.connect(OUT_DB)
    cur = con.cursor()
    cur.execute(f"""
        CREATE TABLE prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {', '.join(f'{c} TEXT' for c in CSV_COLUMNS)}
        )
    """)
    cur.executemany(
        f"INSERT INTO prices ({','.join(CSV_COLUMNS)}) VALUES ({','.join(':'+c for c in CSV_COLUMNS)})",
        rows,
    )
    cur.execute("CREATE INDEX idx_cpt      ON prices(cpt_code)")
    cur.execute("CREATE INDEX idx_payer    ON prices(payer)")
    cur.execute("CREATE INDEX idx_hospital ON prices(hospital_name)")
    con.commit()
    con.close()
    print(f"DB written   → {OUT_DB}  ({os.path.getsize(OUT_DB) / 1024:.0f} KB)")


def print_summary(rows: list[dict]):
    from collections import Counter
    print("\n=== Summary ===")
    for hosp, n in sorted(Counter(r["hospital_name"] for r in rows).items()):
        print(f"  {hosp}: {n:,} rows")
    print()
    for cat, n in sorted(Counter(r["procedure_category"] for r in rows).items()):
        print(f"  {cat}: {n:,} rows")


def main():
    all_rows = []
    for hospital in HOSPITAL_FILES:
        all_rows.extend(parse_hospital_csv(hospital))

    print(f"\nTotal rows across all hospitals: {len(all_rows):,}")
    write_csv(all_rows)
    write_db(all_rows)
    print_summary(all_rows)


if __name__ == "__main__":
    main()
